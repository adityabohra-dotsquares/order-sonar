import csv
from io import StringIO, BytesIO
from typing import List
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from openpyxl import load_workbook
from models.shipping_rules import RateByZone


async def import_rates_by_zone(
    *,
    file_bytes: bytes,
    filename: str,
    db: AsyncSession,
) -> List[RateByZone]:
    ext = filename.split(".")[-1].lower()
    rows = []

    # ---- CSV ----
    if ext == "csv":
        text = file_bytes.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        rows = list(reader)

    # ---- EXCEL ----
    elif ext in ("xlsx", "xls"):
        workbook = load_workbook(BytesIO(file_bytes))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))

    else:
        raise ValueError("Unsupported file format")

    created_rates = []

    for row in rows:
        zone = str(row.get("zone")).strip()
        product_id = str(row.get("product_id")).strip()

        existing = await db.execute(
            select(RateByZone).where(
                RateByZone.zone == zone,
                RateByZone.product_id == product_id,
            )
        )

        if existing.scalar_one_or_none():
            created_rates.append(
                f"{zone} zone and {product_id} product already exists, skipped"
            )
            continue

        rate_obj = RateByZone(
            zone=zone,
            rate=float(row.get("rate", 0)),
            product_id=product_id,
            is_active=str(row.get("is_active", "true")).lower() == "true",
        )

        db.add(rate_obj)
        created_rates.append(f"{zone} zone and {product_id} product added")

    await db.commit()

    return created_rates
